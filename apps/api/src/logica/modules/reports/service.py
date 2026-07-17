"""Servicio de reportes (RF-16, RE-03): exportar el progreso de un grupo a
Excel o PDF. Genera y descarga suceden en request/response separados — la
generación corre como job de arq (nunca bloquea la API, RE-03) y el archivo
queda en disco (`settings.reports_dir`, un volumen compartido en Docker
Compose) para que la API lo transmita cuando esté listo."""

import io
import uuid
from pathlib import Path

import structlog
from arq import ArqRedis
from sqlalchemy.ext.asyncio import AsyncSession

from logica.config import get_settings
from logica.core.errors import NotFoundError, PermissionDeniedError
from logica.modules.groups.models import Group
from logica.modules.groups.service import get_group_with_access
from logica.modules.progress.models import AcademicPeriod
from logica.modules.progress.repository import get_academic_period
from logica.modules.reports import repository
from logica.modules.reports.models import ReportFormat, ReportJob
from logica.modules.reports.repository import StudentReportRow
from logica.modules.users.models import User

logger = structlog.get_logger()


async def request_group_report(
    db: AsyncSession,
    arq_pool: ArqRedis,
    teacher: User,
    group_id: uuid.UUID,
    format: ReportFormat,
    period_id: uuid.UUID | None,
) -> ReportJob:
    _, is_teacher_view = await get_group_with_access(db, teacher, group_id)
    if not is_teacher_view:
        raise PermissionDeniedError("Solo un docente o administrador puede exportar reportes")

    if period_id is not None:
        period = await get_academic_period(db, period_id)
        if period is None or period.institution_id != teacher.institution_id:
            raise NotFoundError("Periodo académico no encontrado")

    job = await repository.create_report_job(
        db, teacher.institution_id, teacher.id, group_id, format, period_id
    )
    await db.flush()
    await arq_pool.enqueue_job("generate_group_report_job", str(job.id))
    return job


def _build_xlsx(group: Group, period: AcademicPeriod | None, rows: list[StudentReportRow]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Progreso"
    ws.append([f"Reporte de progreso — {group.name}"])
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    if period is not None:
        ws.append([f"Periodo: {period.name} ({period.start_date} a {period.end_date})"])
    ws.append([])

    headers = [
        "Estudiante",
        "Correo",
        "Práctica: envíos",
        "Práctica: precisión",
        "Evaluaciones presentadas",
        "Promedio evaluaciones",
        "Insignias",
    ]
    ws.append(headers)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    for row in rows:
        accuracy = row.practice_correct / row.practice_total if row.practice_total else None
        ws.append(
            [
                row.full_name,
                row.email,
                row.practice_total,
                accuracy,
                row.evaluations_submitted,
                row.avg_evaluation_score,
                row.badges_count,
            ]
        )

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _build_pdf(group: Group, period: AcademicPeriod | None, rows: list[StudentReportRow]) -> bytes:
    # Imported lazily: WeasyPrint needs system libraries (Pango/GdkPixbuf,
    # see apps/api/Dockerfile) that aren't necessarily present on every dev
    # machine running `pytest` outside Docker — deferring the import keeps
    # every other report/xlsx test collectible without them installed.
    from weasyprint import HTML

    period_html = (
        f"<p>Periodo: {period.name} ({period.start_date} a {period.end_date})</p>"
        if period is not None
        else ""
    )
    rows_html = "".join(
        f"<tr><td>{r.full_name}</td><td>{r.email}</td><td>{r.practice_total}</td>"
        f"<td>{f'{r.practice_correct / r.practice_total:.0%}' if r.practice_total else '—'}</td>"
        f"<td>{r.evaluations_submitted}</td>"
        f"<td>{f'{r.avg_evaluation_score:.2f}' if r.avg_evaluation_score is not None else '—'}</td>"
        f"<td>{r.badges_count}</td></tr>"
        for r in rows
    )
    html = f"""
    <html><head><style>
        body {{ font-family: sans-serif; }}
        table {{ border-collapse: collapse; width: 100%; }}
        th, td {{ border: 1px solid #999; padding: 4px 8px; text-align: left; }}
        th {{ background: #eee; }}
    </style></head>
    <body>
        <h1>Reporte de progreso — {group.name}</h1>
        {period_html}
        <table>
            <tr><th>Estudiante</th><th>Correo</th><th>Envíos</th><th>Precisión</th>
                <th>Evaluaciones</th><th>Promedio</th><th>Insignias</th></tr>
            {rows_html}
        </table>
    </body></html>
    """
    return HTML(string=html).write_pdf()  # type: ignore[no-any-return]


async def generate_group_report(db: AsyncSession, report_job_id: uuid.UUID) -> None:
    """Called by the arq worker (`generate_group_report_job`) — never on a
    request path. Owns its own commits so the job's status is visible to
    polling clients as soon as each stage finishes, independent of how long
    the file-building step takes."""
    job = await repository.get_report_job(db, report_job_id)
    if job is None:
        return

    await repository.mark_processing(db, job)
    await db.commit()

    try:
        group = await db.get(Group, job.group_id)
        if group is None:
            raise NotFoundError("Grupo no encontrado")
        period = await get_academic_period(db, job.period_id) if job.period_id else None

        period_start = period.start_date if period else None
        period_end = period.end_date if period else None
        rows = await repository.student_report_rows(
            db, job.group_id, period_start=period_start, period_end=period_end
        )

        content = (
            _build_xlsx(group, period, rows)
            if job.format == ReportFormat.xlsx
            else _build_pdf(group, period, rows)
        )

        # noqa comments below: this runs in the arq worker, not on a request
        # path — openpyxl/weasyprint above are already fully synchronous, so
        # a few more blocking filesystem calls change nothing here.
        reports_dir = Path(get_settings().reports_dir)
        reports_dir.mkdir(parents=True, exist_ok=True)  # noqa: ASYNC240
        file_path = reports_dir / f"{job.id}.{job.format.value}"
        file_path.write_bytes(content)  # noqa: ASYNC240

        await repository.mark_done(db, job, str(file_path))
        await db.commit()
    except Exception as exc:
        logger.exception("report_generation_failed", report_job_id=str(report_job_id))
        await repository.mark_failed(db, job, str(exc))
        await db.commit()


async def get_report_job_for_user(
    db: AsyncSession, user: User, report_job_id: uuid.UUID
) -> ReportJob:
    job = await repository.get_report_job(db, report_job_id)
    if job is None or job.institution_id != user.institution_id:
        raise NotFoundError("Reporte no encontrado")
    await get_group_with_access(db, user, job.group_id)
    return job


__all__ = [
    "generate_group_report",
    "get_report_job_for_user",
    "request_group_report",
]
